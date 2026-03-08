"""Output formatting — CSV and Excel export with styled headers and metadata."""

from __future__ import annotations

import io
from datetime import datetime, timezone

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# -- Formatting constants ----------------------------------------------------

_FONT_HEADER = Font(name="IBM Plex Mono", bold=True, size=11)
_FONT_DATA = Font(name="IBM Plex Mono", size=11)
_FONT_META_KEY = Font(name="IBM Plex Sans", bold=True, size=11)
_FONT_META_VAL = Font(name="IBM Plex Sans", size=11)
_FILL_ODD = PatternFill(start_color="F7F6F2", end_color="F7F6F2", fill_type="solid")
_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_ALIGN_WRAP = Alignment(vertical="top", wrap_text=True)


def to_csv(df: pd.DataFrame) -> bytes:
    """Return the DataFrame as UTF-8 encoded CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


def to_excel(df: pd.DataFrame, metadata: dict) -> bytes:
    """Return a styled .xlsx file as bytes with Results and Run Info sheets."""
    wb = Workbook()

    # -- Sheet 1: Results -----------------------------------------------------
    ws_results = wb.active
    ws_results.title = "Results"

    # Write header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws_results.cell(row=1, column=col_idx, value=col_name)
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_WRAP

    # Write data rows with alternating fills
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = _FILL_ODD if row_idx % 2 == 0 else _FILL_EVEN
        for col_idx, value in enumerate(row, start=1):
            cell = ws_results.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _FONT_DATA
            cell.fill = fill
            cell.alignment = _ALIGN_WRAP

    # Auto-width columns (capped at 60 chars)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 502)):  # sample first 500 rows
            val = ws_results.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws_results.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

    # -- Sheet 2: Run Info ----------------------------------------------------
    ws_meta = wb.create_sheet(title="Run Info")

    ws_meta.cell(row=1, column=1, value="Field").font = _FONT_HEADER
    ws_meta.cell(row=1, column=2, value="Value").font = _FONT_HEADER

    for row_idx, (key, value) in enumerate(_flatten_metadata(metadata), start=2):
        key_cell = ws_meta.cell(row=row_idx, column=1, value=key)
        key_cell.font = _FONT_META_KEY

        val_cell = ws_meta.cell(row=row_idx, column=2, value=str(value))
        val_cell.font = _FONT_META_VAL

    ws_meta.column_dimensions["A"].width = 25
    ws_meta.column_dimensions["B"].width = 60

    # -- Write to bytes -------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_metadata(module: str, params: dict, row_count: int) -> dict:
    """Build a standard metadata dict for an export run."""
    return {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "module": module,
        "parameters": params,
        "row_count": row_count,
        "model": "claude-sonnet-4-20250514",
    }


def _flatten_metadata(metadata: dict) -> list[tuple[str, str]]:
    """Flatten a metadata dict into a list of (key, value) pairs for display."""
    rows: list[tuple[str, str]] = []
    for key, value in metadata.items():
        if isinstance(value, dict):
            for sub_key, sub_val in value.items():
                rows.append((f"{key}.{sub_key}", sub_val))
        else:
            rows.append((key, value))
    return rows
